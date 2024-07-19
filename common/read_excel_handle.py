import xlrd
import openpyxl

"""
读取excel文件类和方法的封装
场景1： 把接口测试用例的数据 参数化，写到一个excel表格
     1.读取所有excel的数据,代表所有的用例，
       组装成什么的数据类型
        目前只支持第一个sheet工作表，
        需求作业：包括excel所有的sheet工作表的用例
        [
            （一个用例），
            （二个用例），
            （'admin','11234343'）
        ]
        [
            [一个用例]，
            [二个用例]，
            [三个用例]
        ]
        [
            {
                "username":"admin"
                "password":"123456"
            },
            {},
            {}，
        ]
     2.ddt数据驱动,把excel表格每一行的数据转换成用例
"""


class ReadExcel:
    def __init__(self, filepath):
        self.filepath = filepath
        self.excel = xlrd.open_workbook(self.filepath)
        # self.book = openpyxl.load_workbook(self.filepath)

    def read_excel(self):
        sheets = self.excel.sheet_names()  # 获取excel表格所有的sheet工作表,返回的是列表
        print('接口自动化用例文件有{}个sheet工作表'.format(len(sheets)))
        datas = []  # 定义一个空列表，
        for s in range(len(sheets) - 1):  # s=0,1
            sheet = self.excel.sheet_by_index(s)  # 一次获取每个工作表
            rows, cols = sheet.nrows, sheet.ncols
            # print(rows,cols)
            for i in range(1, rows):  # rows=10,取1,2,3,4。。。。9
                value = sheet.row_values(i)
                # print(value)
                datas.append(value)
            # print(datas)
        return datas

    def write_tiqu(self, name, value):
        """把提取的变量写入到最后一个sheet工作表"""
        sheet = self.book['公共变量']
        # sheet.cell(2,1).value=name
        # sheet.cell(2,2).value=value
        '''公共变量里面的变量名称肯定是唯一，
        所以写入的时候，如果存在就更新
        循环 excel的行数
        '''
        rows = sheet.max_row  # 工作表的行数,3
        for i in range(2, rows + 1):  # i=2,3
            print(sheet.cell(i, 1).value)
            # 表示已经存在，存在就更新  token_tiqu
            # 能够走进if的代码，说明存在相同的，更新值之后，退出函数
            if sheet.cell(i, 1).value == name:
                sheet.cell(i, 2).value = value
                self.book.save(self.filepath)
                return
        # 循环结束没有走到if里面去，说明变量文件不存在重复的名称
        sheet.append([name, value])
        self.book.save(self.filepath)

    def read_tiqu(self, name):
        name = name + '_tiqu'
        sheet = self.excel.sheet_by_name('公共变量')
        for i in range(1, sheet.nrows):  # i=1,2,3
            value = sheet.row_values(i)[0]  # 获取每一行第一列的值
            if value == name:
                return sheet.row_values(i)[1]


if __name__ == '__main__':
    E = ReadExcel(filepath=r'../data/testcase.xls')
    caseDatas = E.read_excel()
    print(caseDatas)
    # E.write_tiqu('token_tiqu','xxxxxx'),
    # print(E.read_tiqu('city'))
