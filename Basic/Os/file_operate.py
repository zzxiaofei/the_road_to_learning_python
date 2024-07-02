import xlwt
import xlrd
import openpyxl
from openpyxl.cell import cell

"""
流程 1.open('路径'，’操作文件的模式’) 2. write 3. close
w:会覆盖 / a：原有内容新增 +的区别是可读可写，注意鼠标的指针位置
"""

"""
w: 文件操作w模式，覆盖写入，清空原文件从头开始写内容，指针到最后
w+: 写+读
"""

# f = open("./openpyxl.xls", mode='w+', encoding='utf-8')
# f.write("123")
# f.seek(0)
# print(f.read())
# f.close()


"""
a：文件操作a模式，从尾方式追加写，不清空不可读
a+: 追加+读
"""
# f = open("test.txt", "a+", encoding="utf-8")
# f.write("\n第四行数据")
# f.seek(0)
# print(f.read())
# # tell = f.tell()
# # print(tell)


"""
r: 文件操作r模式，只读，文件不存在会报错且不清空文件，注意文件指针默认在最前面
大文件不推荐用read(), readlines()，占内存，推荐用readline()，或者循环的方式
r+：读+写
"""
# f = open("test.txt", "r+", encoding="utf-8")
# f.write('345')
# res1 = f.read()
# # res = f.readline()
# print(res1)
# tell = f.tell()  # 获取文件指针位置
# print(tell)
# f.seek(0)  # 移动文件读取指针道指定位置
# tell = f.tell()
# print(tell)


# res_duo = f.readlines()
# print(res1)
# print(res_duo)

# 循环方式读取每行数据
# for line in f.readlines():
#     print(line)

# with open 与 open 区别在于 with open会自动关闭文件
# with open('test.txt', 'r') as f:
#     print(f.read())


"""
xlwt: excel写入，覆盖写入
"""
# book = xlwt.Workbook(encoding='utf-8')  # 1.新建一个excel对象
# sheet_test = book.add_sheet('test1')  # 2.添加一个sheet工作表
# # sheet_test.write(0, 0, 'username')  # 3.单元格行和列分别从0开始


"""
写入一行数据，把数据放到一个列表
"""
# l = ['aaa', 'bbb', 'ccc', 'ddd', 'eee']
# for i in range(0, len(l)):
#     sheet_test.write(0, i, l[i])
#
# book.save('test.xls')  # 4. 保存文件，.save(文件名.xls)

"""
写入一列数据
"""
# col_data = ['qaz', 'wsx', 'edc', 'rfv']
# for data in range(0, len(col_data)):
#     print(data)
#     sheet_test.write(data, 0, col_data[data])
# book.save('text.xls')

"""
xlrd 专注与读取和解析xls
"""

#
# excel = xlrd.open_workbook(filename='openpyxl.xls')
# sheet_test = excel.sheet_by_index(0)  # 指定sheet 从0 开始
# sheet_all = excel.sheet_names()  # 获取所有sheet名称
# print(sheet_all)
# print(sheet_test.nrows, sheet_test.ncols)  # 获取总行数、 总列数

# print(sheet_test.row_values(rowx=0))
# print(sheet_test.cell_value(0, 0))


"""
openpyxl: 读取和写入Excel文件
"""

# excel = openpyxl.load_workbook('openpyxl.xlsx')  # 打开文件
# # sheet_test = excel.active
# # # sheet_test = excel.get_sheet_by_name('openpyxl')
# sheet_test = excel['test1']  # 适用sheet工作表
#
# # 写入数据 单元格写入
# # sheet_test['F4'] = 'Daniel'
# # sheet_test.cell(2, 3).value = 'Pass'
#
# # 整行写入
# new_row = ['post-xml', 'post', 'https://wwww.baidu.com']
# target_row = 1
# for col_num, value in enumerate(new_row, start=1):
#     print(f'{col_num}: {value}')
#     sheet_test.cell(row=target_row, column=col_num, value=value)
#
# # sheet_test.append(new_row)
# excel.save('openpyxl.xlsx')

# c1 = sheet_test.cell(row=1, column=1).value
# print(c1)

# example_list = ['apple', 'banana', 'cherry']
#
# for index, value in enumerate(example_list):
#     print(f'{index}: {value}')


"""
1.homework:写入一列数据['a','b','c',]

写入一行数据
"""
# homework = openpyxl.load_workbook('openpyxl.xlsx')
# sheet = homework.active
# row = ['a', 'b', 'c', 'd', 'e', 'f']
# for col_num in range(1, len(row) + 1):
#     print(col_num)
#     sheet.cell(row=1, column=col_num, value=row[col_num - 1])
# homework.save('openpyxl.xlsx')

# homework = xlwt.Workbook(encodin®g='utf-8')
# worksheet = homework.add_sheet('sheet1')
# list_info = ['a', 'b', 'c', 'd', 'e', 'f']
# for i in range(0, len(list_info)):
#     print(i)
#     worksheet.write(0, i, list_info[i])
# homework.save('./homework.xls')

# [
#     {
#     "username":"admin",
#     "password":"123456",
#     "phone":"1638749385094",
#      },
# ]


homework = xlrd.open_workbook(filename='test.xls')
sheet_test = homework.sheet_by_index(0)  # 指定sheet 从0 开始
header = sheet_test.row_values(0)
num_rows = sheet_test.nrows
num_cols = sheet_test.ncols
# print(num_rows)
# print(num_cols)

list_info = []
for i in range(1, num_rows):
    # print(i)
    data = {}  # i =1,2 j=0,1
    for j in range(num_cols):
        print(j)
        data[header[j]] = sheet_test.row_values(i)[j]
    list_info.append(data)

print(list_info)

# for i in range(num_rows):
#     row_list = []
#     for j in range(num_cols):
#         cell.value = sheet_test.cell_value(i, j)
#         row_list.append(cell.value)
#     print(row_list)
